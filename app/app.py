import streamlit as st
import pandas as pd
import os
import io
import re
from io import BytesIO
from datetime import datetime
from typing import Optional

# ---------------------------
# PAGE CONFIG
# ---------------------------
st.set_page_config(page_title="Excel Utilities — All-in-One", layout="centered")

# ---------------------------
# HELPERS
# ---------------------------
def read_uploaded_bytes(uploaded_file) -> Optional[bytes]:
    if uploaded_file is None:
        return None
    try:
        uploaded_file.seek(0)
    except Exception:
        pass
    return uploaded_file.read()

def to_excel_bytes_from_sheets(sheets: dict) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    output.seek(0)
    return output.getvalue()

# ---------------------------
# DOT/COMMA CONVERTER
# ---------------------------
def swap_dot_comma_df(df: pd.DataFrame) -> pd.DataFrame:
    def swap_cell(x):
        if isinstance(x, str):
            return x.replace('.', 'DOT').replace(',', '.').replace('DOT', ',')
        return x
    return df.applymap(swap_cell)

# ---------------------------
# CSV IKLAN → EXCEL BERWARNA HELPERS
# (adapted from your script)
# ---------------------------
@st.cache_data
def load_uploaded_csv_bytes(file_bytes: bytes) -> pd.DataFrame:
    if file_bytes is None:
        raise ValueError("No file bytes provided")
    raw = file_bytes.decode("utf-8", errors="ignore")
    lines = raw.splitlines()

    HEADER_KEYS = ["Nama Iklan", "Nama Iklan/Produk"]
    header_idx = None
    for i, line in enumerate(lines[:30]):
        if any(k in line for k in HEADER_KEYS):
            header_idx = i
            break
    if header_idx is None:
        raise ValueError("Header Nama Iklan tidak ditemukan")

    delimiter = ";" if lines[header_idx].count(";") > lines[header_idx].count(",") else ","
    clean_csv = "\n".join(lines[header_idx:])
    df = pd.read_csv(io.StringIO(clean_csv), sep=delimiter, engine="python", on_bad_lines="skip")
    df.columns = df.columns.str.strip()
    return df

def normalize_nama_iklan_column(df: pd.DataFrame) -> pd.DataFrame:
    for col in ["Nama Iklan", "Nama Iklan/Produk"]:
        if col in df.columns:
            return df.rename(columns={col: "Nama Iklan"})
    raise ValueError("Kolom Nama Iklan tidak ditemukan")

def short_nama_iklan(nama):
    if pd.isna(nama):
        return nama
    text = str(nama).strip()
    if text.lower().startswith("grup iklan"):
        return text.split(" - ")[0]
    text = re.sub(r"\[.*?\]", "", text).strip()

    feature_blacklist = {
        "busui","friendly","bahan","soft","ultimate","ultimates",
        "motif","size","ukuran","promo","diskon","broad","testing",
        "rayon","katun","cotton","silk","sustra","viscose",
        "linen","polyester","jersey","crepe","chiffon",
        "woolpeach","baloteli","babyterry",
        "pink","hitam","black","putih","white","navy","biru","blue",
        "merah","red","hijau","green","coklat","brown",
        "abu","abu-abu","grey","gray","cream","krem","beige",
        "maroon","ungu","purple","tosca","olive","sage"
    }

    store_blacklist = {
        "official","shop","store","boutique","fashion",
        "my","zahir","myzahir","by","original","premium"
    }

    category_keywords = {
        "gamis","dress","tunik","abaya","set",
        "blouse","khimar","rok","pashmina","hijab","outer",
    }

    context_blacklist = {
        "terbaru","new","update","launch","launching",
        "viral","hits","best","seller","bestseller",
        "kondangan","lebaran","ramadhan","ramadan",
        "harian","pesta","formal","casual",
        "trend","trending","populer",
        "2024","2025","2026","2027", "2028", "2029", "2030"
    }

    parts = re.split(r"\s*[-|]\s*", text)
    product_keywords = {"dress", "gamis", "set"}
    product_candidates = []

    for part in parts:
        words = part.split()
        words_lower = [w.lower() for w in words]
        if not any(w in product_keywords for w in words_lower):
            continue
        while words_lower and words_lower[0] in store_blacklist:
            words_lower.pop(0)
            words.pop(0)
        unique_words = [
            w for w in words_lower
            if w not in store_blacklist
            and w not in feature_blacklist
            and w not in context_blacklist
            and w not in category_keywords
        ]
        if unique_words:
            product_candidates.append(words)

    if product_candidates:
        best_words = product_candidates[-1]
        return " ".join(best_words[:3])

    def score(part):
        s = 0
        for w in part.lower().split():
            if w in store_blacklist:
                s -= 3
            elif w in feature_blacklist:
                s -= 1
            elif w in context_blacklist:
                s -= 2
            elif w in category_keywords:
                s += 1
            else:
                s += 3
        return s

    best = max(parts, key=score)
    return " ".join(best.split()[:3])

def highlight_row(row):
    styles = [''] * len(row)
    roas = row.get('Efektifitas Iklan')
    sales = row.get('Produk Terjual')
    gmv = row.get('Penjualan Langsung (GMV Langsung)')
    cost = row.get('Biaya')

    if pd.isna(sales) or pd.isna(cost):
        return styles

    if (cost == 0) and (sales > 0):
        return ['color: #006400'] * len(row)

    if sales == 0 and cost >= 10000:
        return ['color: #FF0000'] * len(row)

    if sales == 0 and cost < 10000:
        return styles

    if pd.notna(roas):
        try:
            if roas < 8:
                styles = ['background-color: red'] * len(row)
            elif roas < 10:
                styles = ['background-color: yellow'] * len(row)
            else:
                styles = ['background-color: lightgreen'] * len(row)
        except Exception:
            pass

    try:
        nama_idx = row.index.get_loc('Nama Iklan')
    except Exception:
        nama_idx = None
    try:
        gmv_idx = row.index.get_loc('Penjualan Langsung (GMV Langsung)')
    except Exception:
        gmv_idx = None

    if sales > 0 and (pd.isna(gmv) or gmv == 0):
        if nama_idx is not None:
            styles[nama_idx] = 'background-color: lightblue'
        if gmv_idx is not None:
            styles[gmv_idx] = 'background-color: lightblue'
    return styles

def get_iklan_color(row, csv_mode):
    roas = row.get('Efektifitas Iklan')
    sales = row.get('Produk Terjual')
    cost = row.get('Biaya')

    if pd.isna(sales) or pd.isna(cost):
        return None

    if (cost == 0) and (sales > 0):
        return None

    if sales == 0 and cost >= 10000:
        return None

    if sales == 0 and cost < 10000:
        return None

    if csv_mode == "CSV Grup Iklan (hanya iklan produk)":
        if pd.isna(roas):
            return "HIJAU" if sales > 0 else None

    if pd.isna(roas) or roas < 8:
        return "MERAH"
    elif roas < 10:
        return "KUNING"
    else:
        return "HIJAU"

# ---------------------------
# UI — Sidebar Navigation + Coloring Filter
# ---------------------------
st.title("📁 Excel Utilities — Dot/Comma • Sort • Filter • CSV Iklan")
st.write("Pilih fitur di sidebar")

with st.sidebar:
    st.header("Navigation")
    app_mode = st.radio(
        "Pilih fitur",
        options=[
            "Dot ↔ Comma Converter",
            "Sort Penjualan Produk",
            "Filter Nama Produk (Terjual & ATC)",
            "CSV Iklan → Excel Berwarna"
        ]
    )

    # Coloring filter controls (only shown/used by CSV Iklan page)
    st.markdown("---")
    st.subheader("Coloring filter (CSV Iklan)")
    csv_mode_sidebar = st.selectbox(
        "Mode CSV",
        options=["CSV Keseluruhan (Normal)", "CSV Grup Iklan (hanya iklan produk)"],
        index=0
    )
    st.markdown("Pilih kategori yang ingin disertakan di **RINGKASAN_IKLAN** (untuk preview & export)")
    include_merah = st.checkbox("Sertakan MERAH", value=True)
    include_kuning = st.checkbox("Sertakan KUNING", value=True)
    include_hijau = st.checkbox("Sertakan HIJAU", value=True)
    include_biru = st.checkbox("Sertakan BIRU", value=True)
    st.markdown("---")
    st.caption("Catatan: coloring filter hanya mempengaruhi sheet RINGKASAN_IKLAN (preview & export).")

# ---------------------------
# Page: Dot/Comma Converter
# ---------------------------
if app_mode == "Dot ↔ Comma Converter":
    st.header("🔁 Excel Dot ↔ Comma Swapper")
    st.write("Upload file Excel (semua sheet akan diproses). Semua nilai string akan ditukar `.` ↔ `,`.")

    uploaded = st.file_uploader("📂 Upload file Excel (.xlsx/.xls)", type=["xlsx", "xls"], key="dot_uploader")
    if uploaded:
        data = read_uploaded_bytes(uploaded)
        try:
            xls = pd.ExcelFile(BytesIO(data))
            sheets_out = {}
            for sheet_name in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet_name, dtype=str)
                df = swap_dot_comma_df(df)
                sheets_out[sheet_name] = df

            name, ext = os.path.splitext(uploaded.name)
            out_name = f"{name}_dotcomma_swapped.xlsx"
            excel_bytes = to_excel_bytes_from_sheets(sheets_out)

            st.success("✅ File berhasil diproses!")
            st.download_button(
                label="⬇️ Download File Excel (titik-koma tertukar)",
                data=excel_bytes,
                file_name=out_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except Exception as e:
            st.error(f"❌ Terjadi error saat membaca/menulis Excel: {e}")

# ---------------------------
# Page: Sort Penjualan Produk
# ---------------------------
elif app_mode == "Sort Penjualan Produk":
    st.header("📊 Sort Penjualan Produk")
    st.write("Upload file Excel → otomatis di-sort berdasarkan `Channel` lalu `Kode Produk` pada sheet `Performa Produk` (fallback ke sheet pertama jika tidak ada).")

    uploaded = st.file_uploader("Upload file Excel (.xlsx/.xls)", type=["xlsx", "xls"], key="sort_uploader")
    if uploaded:
        data = read_uploaded_bytes(uploaded)
        try:
            xls = pd.ExcelFile(BytesIO(data))
            target_sheet = "Performa Produk" if "Performa Produk" in xls.sheet_names else xls.sheet_names[0]
            df = pd.read_excel(xls, sheet_name=target_sheet)

            st.success(f"File berhasil dibaca (sheet: {target_sheet})")

            required_cols = ["Channel", "Kode Produk"]
            missing = [c for c in required_cols if c not in df.columns]
            if missing:
                st.error(f"Kolom yang diperlukan tidak ditemukan di sheet `{target_sheet}`: {missing}")
            else:
                df_sorted = df.sort_values(by=["Channel", "Kode Produk"], ascending=[True, True])
                st.subheader("Preview Data (20 baris teratas)")
                st.dataframe(df_sorted.head(20), use_container_width=True)

                output = BytesIO()
                df_sorted.to_excel(output, index=False)
                output.seek(0)

                st.download_button(
                    label="⬇️ Download hasil Excel (penjualan_sorted.xlsx)",
                    data=output.getvalue(),
                    file_name="penjualan_sorted.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
        except Exception as e:
            st.error(f"❌ Terjadi error: {e}")

# ---------------------------
# Page: Filter Nama Produk
# ---------------------------
elif app_mode == "Filter Nama Produk (Terjual & ATC)":
    st.header("🧾 Filter Nama Produk (Terjual & ATC)")
    st.write("Upload Excel → ambil nama produk saja → download hasil")

    uploaded = st.file_uploader("Upload file Excel (1 sheet)", type=["xlsx", "xls"], key="filter_uploader")
    if uploaded:
        try:
            df = pd.read_excel(uploaded)
            st.success("File berhasil dibaca")

            required_cols = [
                "Channel",
                "Produk",
                "Produk.1",
                "Produk Ditambahkan ke Keranjang"
            ]
            missing = [c for c in required_cols if c not in df.columns]
            if missing:
                st.error(f"Kolom tidak ditemukan: {missing}")
            else:
                df["Produk.1"] = pd.to_numeric(df["Produk.1"], errors="coerce").fillna(0)
                df["Produk Ditambahkan ke Keranjang"] = pd.to_numeric(df["Produk Ditambahkan ke Keranjang"], errors="coerce").fillna(0)

                df_terjual = (
                    df[df["Produk.1"] > 0][["Channel", "Produk"]]
                    .drop_duplicates()
                    .sort_values(by=["Channel", "Produk"])
                    .reset_index(drop=True)
                )

                df_atc = (
                    df[df["Produk Ditambahkan ke Keranjang"] > 0][["Channel", "Produk"]]
                    .drop_duplicates()
                    .sort_values(by=["Channel", "Produk"])
                    .reset_index(drop=True)
                )

                st.subheader("Preview – Produk Terjual")
                st.dataframe(df_terjual.head(20), use_container_width=True)

                st.subheader("Preview – Produk ATC")
                st.dataframe(df_atc.head(20), use_container_width=True)

                sheets_out = {
                    "Produk Terjual": df_terjual,
                    "Nama Produk ATC": df_atc
                }
                excel_bytes = to_excel_bytes_from_sheets(sheets_out)

                st.download_button(
                    label="⬇️ Download Excel Nama Produk (terjual & atc)",
                    data=excel_bytes,
                    file_name="nama_produk_terjual_dan_atc.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
        except Exception as e:
            st.error(f"Terjadi error: {e}")

# ---------------------------
# Page: CSV Iklan → Excel Berwarna
# ---------------------------
else:
    st.header("📊 CSV Iklan → Excel Berwarna")
    st.write("Upload CSV iklan Shopee → otomatis rapi → download Excel laporan")

    uploaded_file = st.file_uploader("Upload file CSV iklan Shopee", type=["csv"], key="csviklan_uploader")
    csv_mode = csv_mode_sidebar  # controlled from sidebar

    if uploaded_file:
        st.write(f"Mode CSV: **{csv_mode}**")
        # Coloring filter preview toggles
        st.write("Color filter (preview & RINGKASAN only):",
                 f"MERAH: {include_merah}, KUNING: {include_kuning}, HIJAU: {include_hijau}, BIRU: {include_biru}")

        if st.button("🚀 Proses & Download Excel", key="process_csviklan"):
            try:
                with st.spinner("Memproses data..."):
                    raw_bytes = read_uploaded_bytes(uploaded_file)
                    df = load_uploaded_csv_bytes(raw_bytes)
                    df = normalize_nama_iklan_column(df)

                    df["IS_AGGREGATE"] = df["Nama Iklan"].astype(str).str.lower().str.match(r'^\s*grup\b')

                    for col in [
                        "Efektifitas Iklan",
                        "Produk Terjual",
                        "Penjualan Langsung (GMV Langsung)",
                        "Biaya"
                    ]:
                        if col in df.columns:
                            df[col] = pd.to_numeric(df[col], errors="coerce")

                    df["IS_HIJAU_TIPE_A"] = (
                        df.get("Biaya").notna() &
                        (df.get("Biaya") == 0) &
                        (df.get("Produk Terjual") > 0)
                    )

                    df["IS_BIRU"] = (
                        (df.get("Produk Terjual", 0) > 0) &
                        (df.get("Penjualan Langsung (GMV Langsung)", 0) == 0)
                    )

                    df["Nama Ringkasan"] = df["Nama Iklan"].where(
                        df["IS_AGGREGATE"],
                        df["Nama Iklan"].apply(short_nama_iklan)
                    )

                    df["Kategori"] = df.apply(lambda row: get_iklan_color(row, csv_mode), axis=1)

                    if csv_mode == "CSV Grup Iklan (hanya iklan produk)":
                        df_nonagg = df[~df["IS_AGGREGATE"]].copy()
                    else:
                        df_nonagg = df.copy()

                    df_nonagg = df_nonagg[~df_nonagg["IS_HIJAU_TIPE_A"]].copy()

                    ordered_for_numbering = []
                    for kat in ["MERAH", "KUNING", "HIJAU"]:
                        for name in df_nonagg[df_nonagg["Kategori"] == kat]["Nama Ringkasan"]:
                            ordered_for_numbering.append({"nama": name, "kategori": kat})
                    for name in df_nonagg[df_nonagg["IS_BIRU"]]["Nama Ringkasan"]:
                        ordered_for_numbering.append({"nama": name, "kategori": "BIRU"})

                    per_col = {"MERAH": [], "KUNING": [], "HIJAU": [], "BIRU": []}
                    if csv_mode == "CSV Keseluruhan (Normal)":
                        for idx, item in enumerate(ordered_for_numbering, start=1):
                            numbered = f"{idx}. {item['nama']}"
                            per_col[item["kategori"]].append(numbered)
                    else:
                        for kat in ["MERAH", "KUNING", "HIJAU"]:
                            names = df_nonagg[df_nonagg["Kategori"] == kat]["Nama Ringkasan"].tolist()
                            per_col[kat] = [f"{n}," for n in names]
                        names_biru = df_nonagg[df_nonagg["IS_BIRU"]]["Nama Ringkasan"].tolist()
                        per_col["BIRU"] = [f"{n}," for n in names_biru]

                    tanpa_konversi_df = (
                        df_nonagg[(df_nonagg.get("Produk Terjual", 0) == 0) & (df_nonagg.get("Biaya", 0) >= 10000)]
                        [["Nama Ringkasan", "Biaya"]]
                        .rename(columns={"Nama Ringkasan": "Nama Iklan"})
                        .sort_values("Biaya", ascending=False)
                    )

                    hijau_cols = ["Nama Ringkasan", "Produk Terjual", "Efektifitas Iklan", "Biaya"]
                    available_cols = [c for c in hijau_cols if c in df.columns]
                    hijau_tipe_a_df = df[(df.get("Biaya").notna()) & (df.get("Biaya") == 0) & (df.get("Produk Terjual", 0) > 0)][available_cols].copy()
                    if "Nama Ringkasan" in hijau_tipe_a_df.columns:
                        hijau_tipe_a_df = hijau_tipe_a_df.rename(columns={"Nama Ringkasan": "Nama Iklan"})

                    # Apply coloring filter: build filtered per_col copy used for RINGKASAN sheet
                    filtered_per_col = {"MERAH": [], "KUNING": [], "HIJAU": [], "BIRU": []}
                    if include_merah:
                        filtered_per_col["MERAH"] = per_col["MERAH"]
                    if include_kuning:
                        filtered_per_col["KUNING"] = per_col["KUNING"]
                    if include_hijau:
                        filtered_per_col["HIJAU"] = per_col["HIJAU"]
                    if include_biru:
                        filtered_per_col["BIRU"] = per_col["BIRU"]

                    # EXPORT
                    buffer = io.BytesIO()
                    original_name = uploaded_file.name
                    base_name = original_name.rsplit(".", 1)[0]
                    filename = f"{base_name}.xlsx"

                    from openpyxl.styles import Font, Alignment
                    from openpyxl.utils import get_column_letter

                    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
                        # DATA_IKLAN — use Styler to apply highlight_row (if pandas supports to_excel for Styler)
                        try:
                            styled = df.style.apply(highlight_row, axis=1)
                            styled.to_excel(writer, sheet_name="DATA_IKLAN", index=False)
                        except Exception:
                            # fallback: write raw dataframe if style fails
                            df.to_excel(writer, sheet_name="DATA_IKLAN", index=False)

                        wb = writer.book
                        if "RINGKASAN_IKLAN" in wb.sheetnames:
                            wb.remove(wb["RINGKASAN_IKLAN"])
                        ws_ring = wb.create_sheet("RINGKASAN_IKLAN")

                        headers = ["MERAH", "KUNING", "HIJAU", "BIRU"]
                        color_map = {
                            "MERAH": "FF0000",
                            "KUNING": "000000",
                            "HIJAU": "00AA00",
                            "BIRU": "0066CC"
                        }

                        for c_idx, h in enumerate(headers, start=1):
                            cell = ws_ring.cell(row=1, column=c_idx, value=h)
                            cell.font = Font(bold=True)

                        # write content depending on mode, but use filtered_per_col for RINGKASAN
                        if csv_mode == "CSV Keseluruhan (Normal)":
                            for c_idx, key in enumerate(headers, start=1):
                                items = filtered_per_col.get(key, [])
                                if items:
                                    text = "\n".join(items)
                                    cell = ws_ring.cell(row=2, column=c_idx, value=text)
                                    cell.font = Font(color=color_map[key])
                                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                                else:
                                    ws_ring.cell(row=2, column=c_idx, value="")
                        else:
                            for c_idx, key in enumerate(headers, start=1):
                                items = filtered_per_col.get(key, [])
                                if items:
                                    joined = " ".join(items)
                                    if not joined.strip().endswith(","):
                                        joined = joined + ","
                                    cell = ws_ring.cell(row=2, column=c_idx, value=joined)
                                    cell.font = Font(color=color_map[key])
                                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                                else:
                                    ws_ring.cell(row=2, column=c_idx, value="")

                        # adjust column widths
                        for i in range(1, 5):
                            col_letter = get_column_letter(i)
                            ws_ring.column_dimensions[col_letter].width = 40

                        # >10K_TANPA_KONVERSI sheet
                        tanpa_konversi_df.to_excel(writer, sheet_name=">10K_TANPA_KONVERSI", index=False)
                        ws_tc = writer.book[">10K_TANPA_KONVERSI"]
                        for r in range(2, ws_tc.max_row + 1):
                            for c in range(1, ws_tc.max_column + 1):
                                cell = ws_tc.cell(row=r, column=c)
                                cell.font = Font(color="FF0000")

                        # SALES_0_BIAYA (HIJAU TIPE A)
                        hijau_tipe_a_df.to_excel(writer, sheet_name="SALES_0_BIAYA", index=False)
                        ws_hi = writer.book["SALES_0_BIAYA"]
                        for r in range(2, ws_hi.max_row + 1):
                            for c in range(1, ws_hi.max_column + 1):
                                cell = ws_hi.cell(row=r, column=c)
                                cell.font = Font(color="006400")

                    buffer.seek(0)

                st.success("Excel laporan siap di-download 👇")
                st.download_button(
                    "⬇️ Download Excel Laporan",
                    buffer,
                    filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            except Exception as e:
                st.error(f"Terjadi error saat memproses file: {e}")
